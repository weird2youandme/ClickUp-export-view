from turtle import delay
import requests
from datetime import datetime
import buildcustom
import ast
import copy
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font


def addCell(ws, data):
    global c
    global r
    global addr
    addr = ws.cell(row=r,column=c).coordinate
    ws[addr] = data
    c += 1 


# api key is provided by clickup
apikey = input("Enter api key starts with pk_ :")
#team is the first value in the url https://app.clickup.com/########/v
team = input('Enter workspace id :')
#view is the last part of the url when viewing a list or a table
view = input("Enter View ID:")

#The name of the xls report to be created. 
file = input("file name:")

if input("would you like Base columns (URL, Space, folder, list, id)?  Y for yes: ").upper() == 'Y':
    extra = True
else:
    extra = False

if input("would you like add TAGS as a column?  Y for yes: ").upper() == 'Y':
    tags = True
else:
    tags = False


# if new custom fields are created, the custom field json should be rebuilt, this can take a while depending on how many custom fields and unique values exist
if input("would you like to refresh custom field metadata(slow) ?  Y for yes: ").upper() == 'Y':
    print('rebuilding custom fields meta data - this can take a minute')
    custom = buildcustom.customlist()
else:
    cf = open('custom_meta.json', "r")
    ct = cf.read().replace('\n', '')
    custom = ast.literal_eval(ct)
    cf.close()

badlists = ['356837640','356837683'] # place items you would like to not report here - Perhaps look up lists etc...
spaces = {}

payload={}                   
headers = {'Authorization': apikey}
print('creating xlsx file '+file)
url = f'https://api.clickup.com/api/v2/team/{team}/space?archived=false'

while True:
    response = requests.request("GET", url, headers=headers, data=payload)
    if response.status_code == 200:
        break
    elif response.status_code == 429:
        delay(2)
    else:
        print('job failed after receiving invalid response from API response ='+str(response.status_code))
        quit()
temp = response.json()

print('building array of spaces')
for space in temp['spaces']: 
    spaces[space['id']] = space['name']

url = "https://api.clickup.com/api/v2/view/"+view

while True:
    response = requests.request("GET", url, headers=headers, data=payload)
    if response.status_code == 200:
        break
    elif response.status_code == 429:
        delay(2)
    else:
        print('job failed after receiving invalid response from API response ='+str(response.status_code))
        quit()
temp = response.json()

if extra:
    rows = ['URL','Space','folder','list','id']

if tags:
    rows.append('tags')

rows.append('name')

print('building column headers ')
firsttime = True
i = -1
IncludeSubtasks = False
if temp['view']['settings']['show_subtasks'] > 1:
    IncludeSubtasks = True


wb = Workbook()
dest_filename = file+'.xlsx'
ws = wb.active
ws.title = 'ClickUp Report'
hdr = {}

r = 1
c = 1

for row in rows:
    addCell(ws, row)

for column in temp['view']['columns']['fields']: 
    if not column['hidden'] :
        i += 1
        if column['field'][:2]=='cf': #custom fields need to pull the name from custom dict
            if column['field'] == 'cf_e3757138-72ab-4c04-8318-e4348e163870':
                addCell(ws, 'Building')
                addCell(ws, 'Facility')
                addCell(ws, 'Customer')
                addCell(ws, 'Customer#')
            else:    
                addCell(ws, custom[column['field'][3:]]['name'])
            hdr[i] = {"from": 2, "id": column['field'][3:], "type":custom[column['field'][3:]]['type'] }
        else:
            addCell(ws, column['field'])
            hdr[i] = {"from": 1, "id": column['field']}

url = "https://api.clickup.com/api/v2/view/"+view+"/task?page="

i = 0

while True:
    print('pulling page '+str(i)+' from view '+view )
    while True:
        response = requests.request("GET", url+str(i), headers=headers, data=payload)
        if response.status_code == 200:
            break
        elif response.status_code == 429:
            delay(2)
        else:
            print('job failed after receiving invalid response from API response ='+str(response.status_code))
            quit()
    temp = response.json()
    if "err" in temp:
        print("error!  " + temp["err"])
        quit()
    tasks = temp['tasks'] # load result into list
    if IncludeSubtasks: # if suptasks are being displayed, pull them out and add them at top level
        templist = []
        for task in tasks:
            ttask = copy.deepcopy(task)
            if ttask['subtasks_count'] > 0:
                del ttask['subtasks']
                templist.append(ttask)
                for sub in task['subtasks']:  
                    sub['name'] = 'Subtask-'+sub['name'].replace(",", "").replace('•','-')  
                    templist.append(sub)
            else:
                templist.append(ttask)
        tasks = copy.deepcopy(templist)
    
    #walk through tasks pulling data and handle types

    for task in tasks:
        r += 1
        c = 1
        if task['list']['id'] not in badlists:
            if extra:
                addCell(ws, task['url'])
                addCell(ws, spaces[task['space']['id']])
                if task['folder']['hidden']:
                    addCell(ws, '')
                else:
                    addCell(ws, task['folder']['name'])

                addCell(ws, task['list']['name'])
                addCell(ws, task['id'])
            if tags:
                t = ''
                for tag in task['tags']:
                    t += tag['name']+' '
                addCell(ws,t)

            addCell(ws, task['name'])
            for t in hdr:
                h = hdr[t]
                if h['from']==1:
                    if h['id'] == 'status':
                        addCell(ws, task['status']['status'])
                        h = task['status']['color'].lstrip('#')
                        ws[addr].fill = PatternFill("solid", start_color=h)
                        rbg = tuple(int(h[i:i+2], 16) for i in (0, 2, 4))
                        if (rbg[0]*0.299 + rbg[1]*0.587 + rbg[2]*0.114) < 186:
                            ws[addr].font = Font(color='FFFFFFFF')
                    elif h['id'] == 'createdBy':
                        addCell(ws, task['creator']['username'])  
                    elif h['id'] == 'assignee':
                        assign = ''
                        if len(str(task['assignees'])) > 0:
                            for assigned in task['assignees']:
                                if assigned['username'] is not None:
                                    assign += assigned['username'] + ' '
                        addCell(ws, assign)
                        
                    elif h['id'] == 'tags':
                        for tag in task['tags']:
                            addCell(ws, tag['name'])
                        
                    elif h['id']=='priority':
                        if task['priority'] is None:
                            addCell(ws, '') 
                        else:
                            addCell(ws, task['priority']['priority'])
                    elif h['id'] == 'dueDate':
                        if task['due_date'] is None:
                            addCell(ws, '') 
                        else:
                            ts = int(task['due_date'])/1000
                            addCell(ws, datetime.utcfromtimestamp(ts))
                    elif h['id'] == 'startDate':
                        if task['start_date'] is None:
                            addCell(ws, '') 
                        else:
                            ts = int(task['start_date'])/1000
                            addCell(ws, datetime.utcfromtimestamp(ts))
                    elif h['id'] == 'dateUpdated':
                        if task['date_updated'] is None:
                            addCell(ws, '') 
                        else:
                            ts = int(task['date_updated'])/1000
                            addCell(ws, datetime.utcfromtimestamp(ts))
                    elif h['id'] == 'dateClosed':
                        if task['date_closed'] is None:
                            addCell(ws, '') 
                        else:
                            ts = int(task['date_closed'])/1000
                            addCell(ws, datetime.utcfromtimestamp(ts))
                    elif h['id'] == 'dateCreated':
                        if task['date_created'] is None:
                            addCell(ws, '') 
                        else:
                            ts = int(task['date_created'])/1000
                            addCell(ws, datetime.utcfromtimestamp(ts))
                    elif h['id'] == 'timeEstimate':
                        if task['time_estimate'] is None:
                            addCell(ws, '') 
                        else:
                            min = int(task['time_estimate'])/60000
                            addCell(ws, str(min))
                    elif h['id'] == 'timeLogged':
                        if task['time_logged'] is None:
                            addCell(ws, '') 
                        else:
                            min = int(task['time_logged'])/60000
                            addCell(ws, str(min))
                    elif h['id'] == 'incompleteCommentCount':
                        url1 = "https://api.clickup.com/api/v2/task/"+task['id']+"/comment"
                        while True:
                            response = requests.request("GET", url1, headers=headers, data=payload)
                            if response.status_code == 200:
                                break
                            elif response.status_code == 429:
                                delay(2)
                            else:
                                print('job failed after receiving invalid response from API response ='+str(response.status_code))
                        temp1 = response.json()
                        ccount = 0
                        for c in temp1['comments']:
                            if 'resolved' in c:
                                if c['resolved'] is True:
                                    ccount += 1 
                        addCell(ws, str(ccount))
                       
                    elif h['id'] == 'latestComment':
                        url1 = "https://api.clickup.com/api/v2/task/"+task['id']+"/comment"
                        while True:
                            response = requests.request("GET", url1, headers=headers, data=payload)
                            if response.status_code == 200:
                                break
                            elif response.status_code == 429:
                                delay(2)
                            else:
                                print('job failed after receiving invalid response from API response ='+str(response.status_code))
                        temp1 = response.json()
                        if len(temp1['comments']) > 0:
                            addCell(ws, temp1['comments'][0]['comment_text'])
                        else:
                            addCell(ws, '')
                    elif h['id'] == 'lists':
                        if task['list'] is None:
                            addCell(ws, '') 
                        else:
                            addCell(ws, task['list']['name'])
                            
                    else:
                        addCell(ws, task[h['id']])
                else:
                    for custom in task['custom_fields']:
                        if custom['id']== h['id']:
                            if 'value' in custom:
                                if custom['type']=='date':
                                    ts = int(int(custom['value']))/1000
                                    addCell(ws, datetime.utcfromtimestamp(ts)) 
                                elif custom['type'] == 'drop_down':
                                    if list(custom['type_config']['options'][custom['value']]['name']).count('|')==3:
                                        fields = custom['type_config']['options'][custom['value']]['name'].split('|')
                                        addCell(ws, fields[0])
                                        addCell(ws, fields[1])
                                        addCell(ws, fields[2])
                                        addCell(ws, fields[3])
                                    else:
                                        addCell(ws, custom['type_config']['options'][custom['value']]['name'])
                                else:
                                    addCell(ws, custom['value'])
                            else:
                                if list(custom['name']).count('|')==3:
                                        addCell(ws, '')
                                        addCell(ws, '')
                                        addCell(ws, '')
                                        addCell(ws, '')      
                                else:
                                    addCell(ws, '')   


                        
    if temp['last_page']:
        print('last page found - closing file')
        wb.save(filename = dest_filename)
        wb.close()
        break

    i += 1

 



